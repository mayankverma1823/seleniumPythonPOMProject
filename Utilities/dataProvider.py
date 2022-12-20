import openpyxl


def get_data(sheetName):
    workbook = openpyxl.load_workbook("..//excel//TestData.xlsx")
    sheet = workbook[sheetName]
    Totalrows = sheet.max_row
    Totalcols = sheet.max_column
    mainList = []

    for i in range(2, Totalrows + 1):
        dataList = []
        for j in range(1, Totalcols + 1):
            data = sheet.cell(row=i, column=j).value
            dataList.insert(j, data)
        mainList.insert(i, dataList)
    return mainList
